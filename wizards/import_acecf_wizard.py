# -*- coding: utf-8 -*-

import base64
import hashlib
import json
import logging
import uuid
from io import BytesIO
from datetime import datetime

from odoo import api, fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("La libreria 'openpyxl' no esta instalada. No se podra importar ACECF. "
                    "Instalela con: pip install openpyxl")


class ImportAcecfWizard(models.TransientModel):
    _name = "import.acecf.wizard"
    _description = "Asistente para Importar Aprobaciones Comerciales e-CF (ACECF)"

    acecf_file = fields.Binary(string="Archivo ACECF (.xlsx)", required=True)
    filename = fields.Char()
    set_name = fields.Char(
        string="Nombre del Set",
        required=True,
        default=lambda self: f"ACECF {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )

    # Opciones de envio
    send_to_api = fields.Boolean(
        string="Enviar a API automaticamente",
        default=False,
        help="Envia los JSON generados automaticamente a la API despues de importar"
    )

    def _get_workbook(self):
        """Retorna el workbook de Excel cargado"""
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("La libreria 'openpyxl' no esta disponible. Instalela con: pip install openpyxl"))

        if not self.acecf_file:
            raise UserError(_("Por favor, cargue el archivo ACECF."))

        try:
            workbook = openpyxl.load_workbook(BytesIO(base64.b64decode(self.acecf_file)), data_only=True)
            return workbook
        except Exception as e:
            raise UserError(_("No se pudo leer el archivo Excel. Verifique que sea un .xlsx valido. Error: %s") % e)

    def _parse_acecf_sheet(self, sheet):
        """
        Parsea la hoja de ACECF del Excel DGII

        Columnas esperadas:
        - Version
        - RNCEmisor
        - eNCF
        - FechaEmision
        - MontoTotal
        - RNCComprador
        - Estado
        - DetalleMotivoRechazo
        - FechaHoraAprobacionComercial
        """
        rows_data = []

        # Obtener encabezados (fila 1)
        headers = {}
        original_headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                normalized_header = str(cell.value).strip().upper().replace(' ', '').replace('_', '')
                headers[normalized_header] = idx
                original_headers[idx] = str(cell.value).strip()

        _logger.info(f"Encabezados encontrados en hoja ACECF: {list(headers.keys())}")

        # Procesar filas de datos (desde fila 2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            try:
                # Construir diccionario con nombres originales de columna
                excel_row_raw = {}
                for col_idx, col_name in original_headers.items():
                    if col_idx <= len(row):
                        cell_value = row[col_idx - 1]

                        # Convertir a string como lo espera el builder
                        if cell_value is None:
                            pass  # Mantener None
                        elif hasattr(cell_value, 'strftime'):
                            # Fechas: convertir a formato DD-MM-YYYY o DD-MM-YYYY HH:MM:SS
                            if hasattr(cell_value, 'hour'):
                                cell_value = cell_value.strftime('%d-%m-%Y %H:%M:%S')
                            else:
                                cell_value = cell_value.strftime('%d-%m-%Y')
                        elif isinstance(cell_value, bool):
                            cell_value = str(cell_value)
                        elif isinstance(cell_value, (int, float)):
                            # Numeros: convertir a string preservando decimales
                            if isinstance(cell_value, float) and cell_value == int(cell_value):
                                cell_value = str(int(cell_value))
                            else:
                                cell_value = str(cell_value)

                        excel_row_raw[col_name] = cell_value

                # Verificar que tenga datos minimos
                encf = excel_row_raw.get('eNCF') or excel_row_raw.get('ENCF')
                if not encf:
                    _logger.warning(f"Fila {row_idx}: Sin eNCF, omitiendo")
                    continue

                # Extraer datos para el modelo
                rows_data.append({
                    'sequence': row_idx,
                    'fila_excel': row_idx,
                    'excel_row_raw': excel_row_raw,
                    'encf': encf,
                    'version': excel_row_raw.get('Version', '1.0'),
                    'rnc_emisor': excel_row_raw.get('RNCEmisor'),
                    'rnc_comprador': excel_row_raw.get('RNCComprador'),
                    'fecha_emision': excel_row_raw.get('FechaEmision'),
                    'monto_total': self._parse_float(excel_row_raw.get('MontoTotal')),
                    'estado_aprobacion': self._parse_estado(excel_row_raw.get('Estado')),
                    'detalle_motivo_rechazo': excel_row_raw.get('DetalleMotivoRechazo'),
                    'fecha_hora_aprobacion': excel_row_raw.get('FechaHoraAprobacionComercial'),
                })

                _logger.debug(f"Fila ACECF {row_idx} parseada: eNCF={encf}")

            except Exception as e:
                _logger.warning(f"Error al parsear fila {row_idx} de ACECF: {e}")
                continue

        return rows_data

    def _parse_float(self, value):
        """Convierte un valor a float de manera segura"""
        if value is None or value == '':
            return 0.0
        try:
            return float(str(value).replace(',', ''))
        except (ValueError, TypeError):
            return 0.0

    def _parse_estado(self, value):
        """Convierte el estado a seleccion"""
        if value is None:
            return '1'
        try:
            estado = str(int(float(value)))
            return estado if estado in ('1', '2') else '1'
        except (ValueError, TypeError):
            return '1'

    def _hash_payload(self, payload_dict):
        """Genera hash SHA256 del payload"""
        payload_bytes = json.dumps(payload_dict, sort_keys=True, ensure_ascii=False).encode('utf-8')
        return hashlib.sha256(payload_bytes).hexdigest()

    def action_import(self):
        """
        Importa el archivo Excel de ACECF y crea los registros
        """
        self.ensure_one()

        from odoo.addons.l10n_do_e_cf_tests.models import acecf_builder

        _logger.info("Procesando archivo ACECF Excel")
        workbook = self._get_workbook()

        # Buscar hoja de ACECF
        acecf_sheet = None
        for sheet_name in workbook.sheetnames:
            sheet_name_upper = sheet_name.upper()
            if 'ACECF' in sheet_name_upper or 'ACEECF' in sheet_name_upper or 'APROBACION' in sheet_name_upper:
                acecf_sheet = workbook[sheet_name]
                _logger.info(f"Hoja ACECF encontrada: {sheet_name}")
                break

        # Si no encuentra hoja especifica, usar la primera
        if not acecf_sheet and workbook.sheetnames:
            acecf_sheet = workbook[workbook.sheetnames[0]]
            _logger.info(f"Usando primera hoja: {workbook.sheetnames[0]}")

        if not acecf_sheet:
            raise UserError(_("No se encontro hoja ACECF en el archivo Excel."))

        # Parsear hoja
        rows_data = self._parse_acecf_sheet(acecf_sheet)

        if not rows_data:
            raise UserError(_("No se encontraron datos ACECF en el archivo."))

        _logger.info(f"Parseadas {len(rows_data)} filas ACECF")

        # Crear el set de aprobaciones
        acecf_set = self.env['acecf.set'].create({
            'name': self.set_name,
            'description': f'Importado desde {self.filename or "archivo Excel"} el {fields.Datetime.now()}'
        })

        id_lote = str(uuid.uuid4())
        cases_created = 0

        # Crear casos individuales
        for row_data in rows_data:
            try:
                # Construir JSON
                acecf_json = acecf_builder.build_acecf_json(row_data['excel_row_raw'])
                hash_input = self._hash_payload(acecf_json)

                # Crear caso
                case_vals = {
                    'acecf_set_id': acecf_set.id,
                    'name': f"ACECF {row_data['encf']}",
                    'sequence': row_data['sequence'],
                    'fila_excel': row_data['fila_excel'],
                    'id_lote': id_lote,
                    'encf': row_data['encf'],
                    'version': row_data['version'],
                    'rnc_emisor': row_data['rnc_emisor'],
                    'rnc_comprador': row_data['rnc_comprador'],
                    'fecha_emision': row_data['fecha_emision'],
                    'monto_total': row_data['monto_total'],
                    'estado_aprobacion': row_data['estado_aprobacion'],
                    'detalle_motivo_rechazo': row_data['detalle_motivo_rechazo'],
                    'fecha_hora_aprobacion': row_data['fecha_hora_aprobacion'],
                    'payload_json': json.dumps(acecf_json, indent=2, ensure_ascii=False),
                    'hash_input': hash_input,
                    'state': 'payload_ready',
                    'api_status': 'pending',
                }

                case = self.env['acecf.case'].create(case_vals)
                cases_created += 1
                _logger.info(f"Caso ACECF creado: {case.name} (ID: {case.id})")

            except Exception as e:
                _logger.error(f"Error al crear caso para fila {row_data['fila_excel']}: {e}")
                continue

        if cases_created == 0:
            acecf_set.unlink()
            raise UserError(_("No se pudo crear ningun caso ACECF."))

        _logger.info(f"Set ACECF creado: {acecf_set.name} con {cases_created} casos")

        # Enviar a API si esta habilitado
        if self.send_to_api:
            acecf_set.action_send_all()

        # Abrir el set creado
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'acecf.set',
            'res_id': acecf_set.id,
            'view_mode': 'form',
            'target': 'current',
        }
