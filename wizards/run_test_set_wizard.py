import base64
import csv
import hashlib
import json
import logging
import uuid
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from datetime import datetime, date

import requests
from odoo import api, fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
    try:
        from openpyxl.utils.datetime import from_excel as xl_from_excel
    except Exception:
        xl_from_excel = None
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("La librería 'openpyxl' no está instalada. No se podrá ejecutar el set de pruebas. "
                    "Instálela con: pip install openpyxl")


class RunTestSetWizard(models.TransientModel):
    _name = "run.test.set.wizard"
    _description = "Asistente para Ejecutar Set de Pruebas e-CF"

    test_set_file = fields.Binary(string="Archivo Set de Pruebas (.xlsx)", required=True)
    filename = fields.Char()
    test_set_name = fields.Char(string="Nombre del Set", required=True, default="Set de Pruebas DGII")

    # Opciones de ejecución
    send_to_api = fields.Boolean(string="Enviar a API", default=False,
                                 help="Si está marcado, se enviarán los JSON generados automáticamente.")

    # Modo batch - envío controlado por tipo
    batch_mode = fields.Boolean(string="Modo por Lotes (Batch)", default=True,
                                help="Permite enviar por tipo de e-CF, evitando pérdida si algo falla")

    # Filtro por tipo de e-CF
    filter_tipo_ecf = fields.Selection([
        ('all', 'Todos los tipos'),
        ('31', 'Solo tipo 31 - Factura Crédito Fiscal'),
        ('32', 'Solo tipo 32 - Factura Consumo'),
        ('33', 'Solo tipo 33 - Nota Débito'),
        ('34', 'Solo tipo 34 - Nota Crédito'),
        ('41', 'Solo tipo 41 - Compras'),
        ('43', 'Solo tipo 43 - Gastos Menores'),
        ('44', 'Solo tipo 44 - Regímenes Especiales'),
        ('45', 'Solo tipo 45 - Gubernamental'),
        ('46', 'Solo tipo 46 - Exportaciones'),
        ('47', 'Solo tipo 47 - Pagos al Exterior'),
    ], string="Filtrar Tipo e-CF", default='all',
       help="Seleccionar qué tipo de documentos procesar/enviar")

    # NOTA: Los campos de API ahora se toman de la configuración global
    # Ver Configuración > Ajustes > e-CF Tests

    def _is_csv_file(self):
        """Detecta si el archivo cargado es CSV basándose en el nombre"""
        return self.filename and self.filename.lower().endswith('.csv')

    def _get_workbook(self):
        """Retorna el workbook de Excel cargado"""
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("La librería 'openpyxl' no está disponible. Instálela con: pip install openpyxl"))

        if not self.test_set_file:
            raise UserError(_("Por favor, cargue el archivo del set de pruebas."))

        try:
            workbook = openpyxl.load_workbook(BytesIO(base64.b64decode(self.test_set_file)), data_only=True)
            return workbook
        except Exception as e:
            raise UserError(_("No se pudo leer el archivo Excel. Verifique que sea un .xlsx válido. Error: %s") % e)

    def _parse_csv_to_sheet(self):
        """
        Parsea un archivo CSV y retorna una estructura similar a una hoja de Excel.
        Retorna una lista de listas donde cada sublista es una fila.
        """
        if not self.test_set_file:
            raise UserError(_("Por favor, cargue el archivo CSV."))

        try:
            # Decodificar el archivo CSV
            csv_data = base64.b64decode(self.test_set_file).decode('utf-8-sig')

            # Leer el CSV
            csv_reader = csv.reader(StringIO(csv_data))
            rows = list(csv_reader)

            _logger.info(f"CSV parseado: {len(rows)} filas, {len(rows[0]) if rows else 0} columnas")

            return rows
        except Exception as e:
            raise UserError(_("No se pudo leer el archivo CSV. Error: %s") % e)

    # Normalizaciones básicas -------------------------------------------------
    def _normalize_string(self, value):
        if value is None:
            return None
        return str(value).strip()

    def _to_decimal(self, value):
        if value is None or value == '':
            return Decimal('0.00')
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError, TypeError):
            return Decimal('0.00')

    def _normalize_date(self, value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and xl_from_excel:
            try:
                return xl_from_excel(value).date()
            except Exception:
                pass
        if isinstance(value, str):
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return False

    def _hash_payload(self, payload_dict):
        payload_bytes = json.dumps(payload_dict, sort_keys=True, ensure_ascii=False).encode('utf-8')
        return hashlib.sha256(payload_bytes).hexdigest()

    def _extract_payments(self, row, headers):
        payments = []
        for i in range(1, 8):
            forma = self._get_cell_value(row, headers, [f'FORMAPAGO[{i}]', f'FORMAPAGO{i}', f'FORMAPAGO {i}'])
            monto = self._parse_float(self._get_cell_value(row, headers, [f'MONTOPAGO[{i}]', f'MONTOPAGO{i}', f'MONTOPAGO {i}']))
            if forma or monto:
                payments.append({
                    "forma_pago": self._normalize_string(forma) or "",
                    "monto_pago": float(monto),
                })
        return payments

    def _extract_items(self, row, headers, case_data):
        items = []
        for i in range(1, 12):  # El Excel DGII tiene hasta 11+ líneas
            numero = self._get_cell_value(row, headers, [f'NUMEROLINEA[{i}]', f'NUMEROLINEA{i}'])
            nombre = self._get_cell_value(row, headers, [f'NOMBREITEM[{i}]', f'NOMBREITEM{i}'])
            descripcion = self._get_cell_value(row, headers, [f'DESCRIPCIONITEM[{i}]', f'DESCRIPCIONITEM{i}'])
            cantidad = self._parse_float(self._get_cell_value(row, headers, [f'CANTIDADITEM[{i}]', f'CANTIDADITEM{i}']))
            precio_unitario = self._parse_float(self._get_cell_value(row, headers, [f'PRECIOUNITARIOITEM[{i}]', f'PRECIOUNITARIOITEM{i}']))
            monto_item = self._parse_float(self._get_cell_value(row, headers, [f'MONTOITEM[{i}]', f'MONTOITEM{i}']))

            # Campos adicionales DGII
            indicador_facturacion = self._get_cell_value(row, headers, [f'INDICADORFACTURACION[{i}]', f'INDICADORFACTURACION{i}'])
            indicador_bien_servicio = self._get_cell_value(row, headers, [f'INDICADORBIENOSERVICIO[{i}]', f'INDICADORBIENOSERVICIO{i}'])
            unidad_medida = self._get_cell_value(row, headers, [f'UNIDADMEDIDA[{i}]', f'UNIDADMEDIDA{i}'])
            itbis = self._parse_float(self._get_cell_value(row, headers, [f'ITBIS{i}', f'ITBIS[{i}]', f'TOTALITBIS{i}', f'MONTOITBIS[{i}]']))

            # Solo agregar si tiene descripción o nombre, o si tiene monto
            if nombre or descripcion or monto_item or precio_unitario or cantidad:
                item_data = {
                    "linea": int(numero) if numero else i,
                    "cantidad": float(cantidad or 0),
                    "descripcion": self._normalize_string(nombre or descripcion) or f"Item {i}",
                    "precio_unitario": float(precio_unitario or 0),
                    "monto_item": float(monto_item or 0),
                    "itbis": float(itbis or 0),
                }

                # Agregar campos opcionales solo si existen
                if indicador_facturacion:
                    item_data["indicador_facturacion"] = int(indicador_facturacion)
                if indicador_bien_servicio:
                    item_data["indicador_bien_servicio"] = int(indicador_bien_servicio)
                if unidad_medida:
                    item_data["unidad_medida"] = str(unidad_medida)

                items.append(item_data)

        # Si no se encontraron items, crear uno básico
        if not items:
            items.append({
                "linea": 1,
                "cantidad": float(case_data.get('cantidad_items') or 1),
                "descripcion": self._normalize_string(case_data.get('descripcion_item')) or f"Prueba e-CF {case_data.get('tipo_ecf')}",
                "precio_unitario": float(self._to_decimal(case_data.get('precio_unitario') or case_data.get('monto_subtotal'))),
                "monto_item": float(self._to_decimal(case_data.get('monto_total') or case_data.get('monto_subtotal'))),
                "itbis": float(self._to_decimal(case_data.get('total_itbis'))),
                "indicador_bien_servicio": 1,
                "unidad_medida": "Unidad",
            })
        return items

    def _extract_impuestos_adicionales(self, row, headers):
        impuestos = []
        for i in range(1, 5):
            tipo = self._get_cell_value(row, headers, [f'TIPOIMPUESTO[{i}]', f'TIPOIMPUESTO{i}'])
            if tipo:
                impuestos.append({
                    "tipo_impuesto": self._normalize_string(tipo),
                    "isc_especifico": float(self._parse_float(self._get_cell_value(row, headers, [f'MONTOIMPUESTOSELECTIVOCONSUMOESPECIFICO[{i}]', f'MONTOIMPUESTOSELECTIVOCONSUMOESPECIFICO{i}']))),
                    "isc_advalorem": float(self._parse_float(self._get_cell_value(row, headers, [f'MONTOIMPUESTOSELECTIVOCONSUMOADVALOREM[{i}]', f'MONTOIMPUESTOSELECTIVOCONSUMOADVALOREM{i}']))),
                    "otros_impuestos_adicionales": float(self._parse_float(self._get_cell_value(row, headers, [f'OTROSIMPUESTOSADICIONALES[{i}]', f'OTROSIMPUESTOSADICIONALES{i}']))),
                })
        return impuestos

    def _validate_case_data(self, case_data):
        """
        Valida los datos del caso.
        Retorna una tupla (es_valido, lista_de_errores)

        IMPORTANTE: Esta validación es PERMISIVA para permitir que el builder del script
        genere el JSON. El script probado al 100% ya maneja los casos límite.
        Solo validamos que exista tipo_ecf que es imprescindible.
        """
        errors = []

        # Defaults para evitar errores de cabecera incompleta
        if not case_data.get('moneda'):
            case_data['moneda'] = self.env.company.currency_id.name or "DOP"

        if not case_data.get('tipo_ecf'):
            errors.append(_("Falta tipo_ecf"))

        # NOTA: Removemos validaciones estrictas porque el builder del script
        # probado al 100% maneja los valores vacíos/nulos correctamente.
        # La validación real ocurre en DGII al momento del envío.

        # Default para NC/ND si falta razón
        tipo = str(case_data.get('tipo_ecf') or '')
        if tipo in ('33', '34') and not case_data.get('razon_modificacion'):
            case_data['razon_modificacion'] = '01'

        return (len(errors) == 0, errors)

    def _build_canonical_payload(self, case_data, id_lote):
        """
        Construye payload JSON usando el builder del script PROBADO AL 100%
        IMPORTANTE: Usa ecf_builder.build_ecf_json() con la fila RAW del Excel
        """
        from odoo.addons.l10n_do_e_cf_tests.models import ecf_builder

        # ====================================================================
        # CRÍTICO: Usar la fila RAW del Excel si está disponible
        # Esto garantiza que el JSON sea IDÉNTICO al que genera el script
        # ====================================================================
        if case_data.get('excel_row_raw'):
            # Usar directamente la fila raw del Excel
            row = case_data['excel_row_raw']
            _logger.info(f"[JSON BUILDER] Usando fila RAW del Excel - {len(row)} columnas")
            _logger.debug(f"[JSON BUILDER] Columnas disponibles: {list(row.keys())}")
        else:
            # Fallback: construir row desde case_data (modo legacy)
            _logger.warning("[JSON BUILDER] No hay fila RAW del Excel, usando modo legacy")
            row = self._build_row_from_case_data(case_data)

        try:
            # Usar el builder del script PROBADO para construir el JSON
            ecf_json = ecf_builder.build_ecf_json(row)

            # El hash se calcula sobre el JSON generado
            hash_input = self._hash_payload(ecf_json)

            _logger.info(f"[JSON BUILDER] JSON generado exitosamente para tipo {case_data.get('tipo_ecf')}")
            return ecf_json, hash_input

        except Exception as e:
            _logger.error(f"[JSON BUILDER] Error al construir JSON: {str(e)}", exc_info=True)
            raise

    def _build_row_from_case_data(self, case_data):
        """
        Fallback: Construye row desde case_data si no hay fila RAW disponible
        Usado solo para compatibilidad con datos antiguos
        """
        row = {}

        # Version y tipo
        row["Version"] = case_data.get('version') or "1.0"
        row["TipoeCF"] = case_data.get('tipo_ecf')
        row["eNCF"] = case_data.get('encf') or ""
        row["ENCF"] = case_data.get('encf') or ""

        # IdDoc
        row["IndicadorNotaCredito"] = case_data.get('indicador_nota_credito')
        row["FechaVencimientoSecuencia"] = case_data.get('fecha_vencimiento_secuencia')
        row["IndicadorMontoGravado"] = case_data.get('indicador_monto_gravado')
        row["TipoIngresos"] = case_data.get('tipo_ingreso') or "01"
        row["TipoPago"] = case_data.get('tipo_pago') or "1"

        # Emisor
        row["RNCEmisor"] = case_data.get('rnc_emisor') or self.env.company.vat or ""
        row["RazonSocialEmisor"] = case_data.get('razon_social_emisor') or self.env.company.name or ""
        row["NombreComercial"] = case_data.get('nombre_comercial') or ""
        row["DireccionEmisor"] = case_data.get('direccion_emisor') or ""
        row["Municipio"] = case_data.get('municipio_emisor') or ""
        row["Provincia"] = case_data.get('provincia_emisor') or ""
        row["CorreoEmisor"] = case_data.get('correo_emisor') or ""
        row["WebSite"] = case_data.get('website_emisor') or ""
        row["CodigoVendedor"] = case_data.get('codigo_vendedor') or ""
        row["NumeroFacturaInterna"] = case_data.get('numero_factura_interna') or ""
        row["NumeroPedidoInterno"] = case_data.get('numero_pedido_interno') or ""
        row["ZonaVenta"] = case_data.get('zona_venta') or ""

        # Fecha de emisión en formato DD-MM-YYYY
        if case_data.get('fecha_comprobante'):
            if hasattr(case_data.get('fecha_comprobante'), 'strftime'):
                row["FechaEmision"] = case_data.get('fecha_comprobante').strftime('%d-%m-%Y')
            else:
                row["FechaEmision"] = str(case_data.get('fecha_comprobante'))
        else:
            row["FechaEmision"] = datetime.now().strftime('%d-%m-%Y')

        # Comprador
        row["IdentificadorExtranjero"] = case_data.get('identificador_extranjero') or ""
        row["RNCComprador"] = case_data.get('receptor_rnc') or ""
        row["RazonSocialComprador"] = case_data.get('receptor_nombre') or ""

        # Totales
        row["MontoGravadoTotal"] = case_data.get('monto_gravado_total') or case_data.get('monto_gravado_i1') or ""
        row["MontoGravadoI1"] = case_data.get('monto_gravado_i1') or ""
        row["MontoGravadoI2"] = case_data.get('monto_gravado_i2') or ""
        row["MontoGravadoI3"] = case_data.get('monto_gravado_i3') or ""
        row["MontoExento"] = case_data.get('monto_exento') or ""
        row["TotalITBIS"] = case_data.get('total_itbis') or ""
        row["TotalITBIS1"] = case_data.get('total_itbis1') or ""
        row["TotalITBIS2"] = case_data.get('total_itbis2') or ""
        row["TotalITBIS3"] = case_data.get('total_itbis3') or ""
        row["MontoTotal"] = case_data.get('monto_total') or 0
        row["ValorPagar"] = case_data.get('monto_total_pagar') or case_data.get('monto_total') or 0
        row["MontoNoFacturable"] = case_data.get('monto_no_facturable') or ""

        # Items básicos
        items = case_data.get('items') or []
        if items:
            for idx, item in enumerate(items, 1):
                row[f"NumeroLinea[{idx}]"] = item.get('linea', idx)
                row[f"NombreItem[{idx}]"] = item.get('descripcion', f"Item {idx}")
                row[f"CantidadItem[{idx}]"] = item.get('cantidad', 1)
                row[f"UnidadMedida[{idx}]"] = item.get('unidad_medida', 'Unidad')
                row[f"PrecioUnitarioItem[{idx}]"] = item.get('precio_unitario', 0)
                row[f"MontoItem[{idx}]"] = item.get('monto_item', 0)
                if item.get('indicador_facturacion'):
                    row[f"IndicadorFacturacion[{idx}]"] = item.get('indicador_facturacion')
                if item.get('indicador_bien_servicio'):
                    row[f"IndicadorBienoServicio[{idx}]"] = item.get('indicador_bien_servicio')
        else:
            row["NumeroLinea[1]"] = 1
            row["NombreItem[1]"] = case_data.get('descripcion_item') or f"Prueba e-CF {case_data.get('tipo_ecf')}"
            row["CantidadItem[1]"] = case_data.get('cantidad_items') or 1
            row["UnidadMedida[1]"] = "Unidad"
            row["PrecioUnitarioItem[1]"] = case_data.get('precio_unitario') or case_data.get('monto_subtotal') or 0
            row["MontoItem[1]"] = case_data.get('monto_total') or 0
            row["IndicadorBienoServicio[1]"] = 1

        # Referencia para NC/ND
        if case_data.get('ncf_modificado'):
            row["NCFModificado"] = case_data.get('ncf_modificado')
            row["CodigoModificacion"] = case_data.get('razon_modificacion') or "01"

        return row

    def _build_rfce_payload(self, case_data, id_lote):
        """Construye payload JSON para RFCE (referencia, no se envía)."""
        pagos = case_data.get('pagos') or []
        impuestos_adicionales = case_data.get('impuestos_adicionales') or []
        payload = {
            "meta": {
                "caso_prueba": self._normalize_string(case_data.get('caso_prueba')) or f"RFCE {case_data.get('sequence')}",
                "version": self._normalize_string(case_data.get('version')) or "1.0",
                "tipo_ecf": 32,
                "encf": self._normalize_string(case_data.get('encf')) or "",
                "codigo_seguridad_ecf": self._normalize_string(case_data.get('codigo_seguridad_ecf')),
                "fila_excel": case_data.get('sequence'),
                "id_lote": id_lote,
                "hash_input": None,
            },
            "encabezado": {
                "tipo_ingresos": self._normalize_string(case_data.get('tipo_ingreso')),
                "tipo_pago": self._normalize_string(case_data.get('tipo_pago')),
            },
            "pagos": pagos,
            "impuestos_adicionales": impuestos_adicionales,
            "totales": {
                "monto_total": float(self._to_decimal(case_data.get('monto_total'))),
                "monto_no_facturable": float(self._to_decimal(case_data.get('monto_no_facturable'))),
                "monto_periodo": case_data.get('monto_periodo'),
            }
        }
        hash_input = self._hash_payload(payload)
        payload["meta"]["hash_input"] = hash_input
        return payload, hash_input

    # ========================================================================
    # FUNCIONES DE MSeller API (del script probado al 100%)
    # ========================================================================

    def _mseller_login(self, host, env, email, password, timeout=30):
        """
        Login a MSeller API para obtener token de autenticación
        Función EXACTA del script probado
        """
        import sys

        url = f"{host.rstrip('/')}/{env}/customer/authentication"
        payload = {"email": email, "password": password}

        try:
            r = requests.post(url, json=payload, timeout=timeout)
        except Exception as e:
            raise UserError(_(
                "Error de conexión al intentar login en MSeller:\n"
                "URL: %s\n"
                "Error: %s"
            ) % (url, str(e)))

        try:
            data = r.json()
        except Exception:
            raise UserError(_(
                "Login MSeller: respuesta no-JSON (%s):\n%s"
            ) % (r.status_code, r.text[:300]))

        if r.status_code >= 400:
            raise UserError(_(
                "Login MSeller falló (%s):\n%s"
            ) % (r.status_code, json.dumps(data, ensure_ascii=False)))

        token = data.get("idToken") or data.get("token") or data.get("accessToken")
        if not token:
            raise UserError(_(
                "Login MSeller OK pero no se encontró token en respuesta:\n%s"
            ) % json.dumps(data, ensure_ascii=False))

        return token

    def _mseller_send_doc(self, host, env, api_key, bearer, doc, timeout=60):
        """
        Envía un documento e-CF a MSeller API
        Función EXACTA del script probado
        Retorna: (status_code, response_json_or_text)
        """
        url = f"{host.rstrip('/')}/{env}/documentos-ecf"

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {bearer}",
            "X-API-KEY": api_key,
        }

        try:
            r = requests.post(url, headers=headers, json=doc, timeout=timeout)
        except Exception as e:
            raise UserError(_(
                "Error al enviar documento a MSeller:\n"
                "URL: %s\n"
                "Error: %s"
            ) % (url, str(e)))

        try:
            return r.status_code, r.json()
        except Exception:
            return r.status_code, r.text

    def _get_tipo_ecf_priority(self, doc):
        """
        Retorna prioridad de envío según TipoeCF (menor = primero)
        Función del script probado para ordenar documentos según requisitos DGII
        """
        try:
            tipo_ecf = doc["ECF"]["Encabezado"]["IdDoc"].get("TipoeCF", "99")

            # Grupo 1 - Primero
            if tipo_ecf == "31":
                return (1, 1)  # Factura Crédito Fiscal
            elif tipo_ecf == "32":
                # Determinar si es >=250k, Resumen o <250k
                monto_total = doc["ECF"]["Encabezado"].get("Totales", {}).get("MontoTotal")
                if monto_total:
                    try:
                        monto = float(str(monto_total).replace(",", ""))
                        if monto >= 250000:
                            return (1, 2)  # 32 Mayor o Igual 250k
                        return (4, 1)  # 32 Menor a 250k
                    except:
                        return (4, 1)
                return (4, 1)
            elif tipo_ecf == "41":
                return (1, 3)  # Compras
            elif tipo_ecf == "43":
                return (1, 4)  # Gastos Menores
            elif tipo_ecf == "44":
                return (1, 5)  # Regímenes Especiales
            elif tipo_ecf == "45":
                return (1, 6)  # Gubernamental
            elif tipo_ecf == "46":
                return (1, 7)  # Exportaciones
            elif tipo_ecf == "47":
                return (1, 8)  # Pagos al Exterior

            # Grupo 2 - Segundo
            elif tipo_ecf == "33":
                return (2, 1)  # Nota de Débito
            elif tipo_ecf == "34":
                return (2, 2)  # Nota de Crédito

            # Otros tipos al final
            else:
                return (99, int(tipo_ecf) if str(tipo_ecf).isdigit() else 99)

        except Exception as e:
            _logger.warning(f"No se pudo determinar prioridad de documento: {e}")
            return (99, 99)

    def _send_to_internal_api(self, payload, test_case=None, id_lote=None, fila_excel=None):
        """
        Envía el payload a la API interna (JSON -> XML + firma + envío).
        Genera un log detallado de cada llamada para trazabilidad completa.
        """
        import time

        # Obtener configuración global
        ICP = self.env["ir.config_parameter"].sudo()
        api_url = ICP.get_param("l10n_do_e_cf_tests.api_url")
        api_token = ICP.get_param("l10n_do_e_cf_tests.api_token")
        timeout = int(ICP.get_param("l10n_do_e_cf_tests.api_timeout", "30"))
        enable_debug_log = ICP.get_param("l10n_do_e_cf_tests.enable_debug_log", "True") == "True"

        if not api_url:
            raise UserError(_(
                "Configure la URL de la API interna en:\n"
                "Configuración > Ajustes > e-CF Tests > URL API e-CF"
            ))

        # Preparar headers
        headers = {"Content-Type": "application/json"}
        if api_token:
            headers["Authorization"] = f"Bearer {api_token}"

        # Crear log inicial
        log_vals = {
            'test_case_id': test_case.id if test_case else False,
            'id_lote': id_lote,
            'fila_excel': fila_excel,
            'request_url': api_url,
            'request_method': 'POST',
            'request_headers': json.dumps(headers, indent=2) if enable_debug_log else None,
            'request_payload': json.dumps(payload, indent=2, ensure_ascii=False) if enable_debug_log else None,
            'request_timestamp': fields.Datetime.now(),
            'api_status': 'pending',
        }

        api_log = self.env['ecf.api.log'].create(log_vals)

        # Variables de respuesta
        track_id = None
        accepted = False
        rejected = False
        resp_text = ""
        error_message = None
        dgii_message = None
        dgii_code = None

        try:
            # Realizar la llamada HTTP
            start_time = time.time()
            resp = requests.post(api_url, json=payload, headers=headers, timeout=timeout)
            end_time = time.time()
            response_time_ms = int((end_time - start_time) * 1000)

            resp_text = resp.text
            status_code = resp.status_code

            # Intentar parsear la respuesta JSON
            try:
                resp_json = resp.json()
                track_id = resp_json.get("track_id") or resp_json.get("trackId")
                status_val = (resp_json.get("status") or resp_json.get("estado") or "").upper()
                dgii_message = resp_json.get("message") or resp_json.get("mensaje")
                dgii_code = resp_json.get("code") or resp_json.get("codigo")

                if status_val in ("ACEPTADO", "ACCEPTED", "OK"):
                    accepted = True
                elif status_val in ("RECHAZADO", "REJECTED", "ERROR"):
                    rejected = True

            except Exception as json_error:
                _logger.warning(f"No se pudo parsear respuesta JSON de API: {json_error}")
                resp_json = None

            # Determinar estado del log
            if resp.ok:
                if accepted:
                    api_status = 'accepted'
                elif rejected:
                    api_status = 'rejected'
                else:
                    api_status = 'success'
            else:
                api_status = 'error'
                error_message = f"HTTP {status_code}: {resp_text[:500]}"

            # Actualizar log con respuesta
            api_log.write({
                'response_status_code': status_code,
                'response_headers': json.dumps(dict(resp.headers), indent=2) if enable_debug_log else None,
                'response_body': resp_text if enable_debug_log else resp_text[:1000],
                'response_timestamp': fields.Datetime.now(),
                'response_time_ms': response_time_ms,
                'track_id': track_id,
                'api_status': api_status,
                'error_message': error_message,
                'dgii_message': dgii_message,
                'dgii_code': dgii_code,
            })

            # Si no es exitoso, lanzar error
            if not resp.ok and not (accepted or rejected):
                raise UserError(_(
                    "API interna devolvió error HTTP %s:\n%s\n\n"
                    "Ver log #%s para más detalles."
                ) % (resp.status_code, resp_text[:200], api_log.id))

            _logger.info(
                f"API call exitoso - Log #{api_log.id} - "
                f"Status: {api_status} - Time: {response_time_ms}ms - "
                f"Track ID: {track_id or 'N/A'}"
            )

        except requests.Timeout:
            error_message = f"Timeout después de {timeout} segundos"
            api_log.write({
                'api_status': 'timeout',
                'error_message': error_message,
                'response_timestamp': fields.Datetime.now(),
            })
            _logger.error(f"API Timeout - Log #{api_log.id}: {error_message}")
            raise UserError(_("La API no respondió en %s segundos.\nVer log #%s") % (timeout, api_log.id))

        except requests.ConnectionError as conn_error:
            error_message = f"Error de conexión: {str(conn_error)}"
            api_log.write({
                'api_status': 'connection_error',
                'error_message': error_message,
                'response_timestamp': fields.Datetime.now(),
            })
            _logger.error(f"API Connection Error - Log #{api_log.id}: {error_message}")
            raise UserError(_(
                "No se pudo conectar a la API en: %s\n"
                "Error: %s\n\n"
                "Verifique la URL en Configuración > Ajustes > e-CF Tests\n"
                "Ver log #%s"
            ) % (api_url, str(conn_error)[:200], api_log.id))

        except Exception as e:
            error_message = f"Error inesperado: {str(e)}"
            api_log.write({
                'api_status': 'error',
                'error_message': error_message,
                'response_timestamp': fields.Datetime.now(),
            })
            _logger.error(f"API Unexpected Error - Log #{api_log.id}: {error_message}", exc_info=True)
            raise UserError(_("Error inesperado al llamar la API:\n%s\n\nVer log #%s") % (str(e), api_log.id))

        return resp_text, track_id, accepted, rejected, api_log

    def _parse_ecf_csv(self, csv_rows):
        """
        Parsea un archivo CSV con datos de prueba DGII.
        csv_rows es una lista de listas (cada sublista es una fila).
        """
        cases = []

        if not csv_rows or len(csv_rows) < 2:
            _logger.warning("CSV vacío o sin datos")
            return cases

        # Primera fila son los encabezados
        header_row = csv_rows[0]
        headers = {}
        for idx, cell_value in enumerate(header_row, start=1):
            if cell_value:
                # Normalizar: eliminar espacios, convertir a mayúsculas, eliminar guiones bajos
                normalized_header = str(cell_value).strip().upper().replace(' ', '').replace('_', '')
                headers[normalized_header] = idx

        _logger.info(f"Encabezados encontrados en CSV: {list(headers.keys())}")

        # Procesar filas de datos (desde fila 2)
        for row_idx, row in enumerate(csv_rows[1:], start=2):
            if not row or not any(row):
                continue

            try:
                # Mapeo flexible de columnas
                tipo_ecf = self._get_cell_value(row, headers, ['TIPOECF', 'TIPOЁCF', 'TIPO', 'TIPO ECF', 'TIPO_ECF'])
                if not tipo_ecf:
                    continue

                case_data = {
                    'caso_prueba': self._get_cell_value(row, headers, ['CASOPRUEBA', 'CASO']),
                    'version': self._get_cell_value(row, headers, ['VERSION']),
                    'encf': self._get_cell_value(row, headers, ['ENCF']),
                    'codigo_seguridad_ecf': self._get_cell_value(row, headers, ['CODIGOSEGURIDADECF', 'CODIGOSEGURIDAD', 'CODSEG']),
                    'sequence': row_idx,
                    'name': f"Caso {row_idx} - Tipo {tipo_ecf}",
                    'tipo_ecf': str(tipo_ecf).strip(),
                    'receptor_rnc': self._get_cell_value(row, headers, ['RNCCOMPRADOR', 'RNCRECEPTOR', 'RNC', 'RECEPTOR RNC', 'RNC RECEPTOR', 'RNC_RECEPTOR']),
                    'receptor_nombre': self._get_cell_value(row, headers, ['RAZONSOCIALCOMPRADOR', 'NOMBRERECEPTOR', 'NOMBRE', 'RECEPTOR', 'NOMBRE RECEPTOR', 'RECEPTOR_NOMBRE']),
                    'fecha_comprobante': self._parse_date(self._get_cell_value(row, headers, ['FECHAEMISION', 'FECHACOMPROBANTE', 'FECHA', 'FECHA COMPROBANTE', 'FECHA_COMPROBANTE'])),
                    'moneda': self._get_cell_value(row, headers, ['MONEDA', 'MONEDAFACTURA', 'CURRENCY', 'TIPOMONEDA']),
                    'tipo_cambio': self._get_cell_value(row, headers, ['TIPOCAMBIO', 'TIPO CAMBIO']),
                    'tipo_ingreso': self._get_cell_value(row, headers, ['TIPOINGRESOS', 'TIPO INGRESO', 'TIPO_INGRESO', 'INGRESOS']),
                    'tipo_pago': self._get_cell_value(row, headers, ['TIPOPAGO', 'FORMAPAGO[1]', 'FORMAPAGO', 'TIPO PAGO', 'TIPO_PAGO', 'PAGO']),
                    'monto_subtotal': self._parse_float(self._get_cell_value(row, headers, ['MONTOSUBTOTAL', 'SUBTOTAL', 'MONTO SUBTOTAL', 'MONTO_SUBTOTAL'])),
                    'monto_descuento': self._parse_float(self._get_cell_value(row, headers, ['MONTODESCUENTO', 'DESCUENTO'])),
                    'monto_gravado_i1': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADOI1', 'MONTOGRAVADOTASAI1', 'GRAVADO I1', 'MONTO GRAVADO I1', 'GRAVADO_I1'])),
                    'monto_gravado_i2': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADOI2', 'MONTOGRAVADOTASAI2', 'GRAVADO I2', 'MONTO GRAVADO I2', 'GRAVADO_I2'])),
                    'monto_gravado_i3': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADOI3', 'MONTOGRAVADOTASAI3', 'GRAVADO I3', 'MONTO GRAVADO I3', 'GRAVADO_I3'])),
                    'monto_exento': self._parse_float(self._get_cell_value(row, headers, ['MONTOEXENTO', 'EXENTO', 'MONTO EXENTO'])),
                    'total_itbis': self._parse_float(self._get_cell_value(row, headers, ['TOTALITBIS', 'ITBIS', 'TOTAL ITBIS', 'ITBIS TOTAL'])),
                    'total_itbis1': self._parse_float(self._get_cell_value(row, headers, ['ITBIS1', 'ITBIS 1'])),
                    'total_itbis2': self._parse_float(self._get_cell_value(row, headers, ['ITBIS2', 'ITBIS 2'])),
                    'total_itbis3': self._parse_float(self._get_cell_value(row, headers, ['ITBIS3', 'ITBIS 3'])),
                    'monto_total': self._parse_float(self._get_cell_value(row, headers, ['MONTOTOTAL', 'TOTAL', 'MONTO TOTAL', 'MONTO_TOTAL'])),
                    'ncf_modificado': self._get_cell_value(row, headers, ['ENCFMODIFICADO', 'NCFMODIFICADO', 'NCF MODIFICADO', 'NCF_MODIFICADO']),
                    'razon_modificacion': self._map_razon_modificacion(self._get_cell_value(row, headers, ['RAZONMODIFICACION', 'RAZON', 'RAZON MODIFICACION', 'RAZON_MODIFICACION', 'CODIGOMODIFICACION'])),
                    'expected_result': self._get_cell_value(row, headers, ['RESULTADOESPERADO', 'RESULTADO', 'ESPERADO', 'RESULTADO ESPERADO']),
                    'identificador_extranjero': self._get_cell_value(row, headers, ['IDENTIFICADOREXTRANJERO', 'IDEXTRANJERO', 'IDENT_EXTRANJERO']),
                    # Emisor
                    'rnc_emisor': self._get_cell_value(row, headers, ['RNCEMISOR']),
                    'razon_social_emisor': self._get_cell_value(row, headers, ['RAZONSOCIALEMISOR']),
                    'nombre_comercial': self._get_cell_value(row, headers, ['NOMBRECOMERCIAL']),
                    'sucursal': self._get_cell_value(row, headers, ['SUCURSAL']),
                    'direccion_emisor': self._get_cell_value(row, headers, ['DIRECCIONEMISOR']),
                    'municipio_emisor': self._get_cell_value(row, headers, ['MUNICIPIO']),
                    'provincia_emisor': self._get_cell_value(row, headers, ['PROVINCIA']),
                    'correo_emisor': self._get_cell_value(row, headers, ['CORREOEMISOR']),
                    'website_emisor': self._get_cell_value(row, headers, ['WEBSITE']),
                    'actividad_economica': self._get_cell_value(row, headers, ['ACTIVIDADECONOMICA']),
                    'codigo_vendedor': self._get_cell_value(row, headers, ['CODIGOVENDEDOR']),
                    'numero_factura_interna': self._get_cell_value(row, headers, ['NUMEROFACTURAINTERNA']),
                    'numero_pedido_interno': self._get_cell_value(row, headers, ['NUMEROPEDIDOINTERNO']),
                    'zona_venta': self._get_cell_value(row, headers, ['ZONAVENTA']),
                    'ruta_venta': self._get_cell_value(row, headers, ['RUTAVENTA']),
                    'info_adicional_emisor': self._get_cell_value(row, headers, ['INFORMACIONADICIONALEMISOR']),
                    'monto_no_facturable': self._parse_float(self._get_cell_value(row, headers, ['MONTONOFACTURABLE'])),
                }

                # Calcular monto total a pagar
                case_data['monto_total_pagar'] = case_data['monto_total']

                # Calcular precio unitario para la línea
                case_data['cantidad_items'] = 1
                case_data['precio_unitario'] = case_data['monto_subtotal'] or 0
                case_data['descripcion_item'] = f"Prueba e-CF Tipo {tipo_ecf}"
                case_data['pagos'] = self._extract_payments(row, headers)
                case_data['items'] = self._extract_items(row, headers, case_data)

                cases.append(case_data)

                _logger.debug(f"Caso ECF parseado desde CSV: {case_data['name']}")

            except Exception as e:
                _logger.warning(f"Error al parsear fila {row_idx} del CSV: {e}")
                continue

        return cases

    def _parse_ecf_sheet(self, sheet):
        """
        Parsea la hoja 'ECF' del Excel DGII
        Estructura esperada (columnas ejemplo):
        A: Secuencia
        B: Tipo e-CF (31, 32, 33, 34, etc.)
        C: RNC/Cédula Receptor
        D: Nombre Receptor
        E: Fecha Comprobante
        F: Tipo Ingresos
        G: Tipo Pago
        H: Monto Subtotal
        I: Descuento
        J: Monto Gravado I1
        K: Monto Gravado I2
        L: Monto Gravado I3
        M: Monto Exento
        N: ITBIS Total
        O: ITBIS 1
        P: ITBIS 2
        Q: ITBIS 3
        R: Total
        S: NCF Modificado
        T: Razón Modificación
        U: Resultado Esperado
        """
        cases = []

        # Obtener encabezados (fila 1)
        headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                # Normalizar: eliminar espacios, convertir a mayúsculas, eliminar guiones bajos
                normalized_header = str(cell.value).strip().upper().replace(' ', '').replace('_', '')
                headers[normalized_header] = idx

        _logger.info(f"Encabezados encontrados en hoja ECF: {list(headers.keys())}")

        # Crear mapeo inverso: índice -> nombre original de columna
        # Esto permite pasar la fila RAW al builder del script
        original_headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                original_headers[idx] = str(cell.value).strip()

        # Procesar filas de datos (desde fila 2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            # Extraer valores según columnas
            try:
                # Mapeo flexible de columnas (soporta nombres DGII camelCase y variaciones)
                tipo_ecf = self._get_cell_value(row, headers, ['TIPOECF', 'TIPOЁCF', 'TIPO', 'TIPO ECF', 'TIPO_ECF'])
                if not tipo_ecf:
                    continue

                # ====================================================================
                # IMPORTANTE: Guardar la fila RAW como diccionario para el builder
                # Esto permite que ecf_builder.build_ecf_json() reciba los datos
                # EXACTAMENTE como el script original los leía del Excel
                #
                # NOTA: pandas.read_excel con dtype=object lee todo como strings.
                # openpyxl devuelve tipos nativos, así que convertimos a string
                # para replicar el comportamiento del script.
                # ====================================================================
                excel_row_raw = {}
                for col_idx, col_name in original_headers.items():
                    if col_idx <= len(row):
                        cell_value = row[col_idx - 1]

                        # Convertir a string como lo hace pandas con dtype=object
                        if cell_value is None:
                            pass  # Mantener None para que is_empty() funcione
                        elif hasattr(cell_value, 'strftime'):
                            # Fechas: convertir a formato DD-MM-YYYY
                            cell_value = cell_value.strftime('%d-%m-%Y')
                        elif isinstance(cell_value, bool):
                            cell_value = str(cell_value)
                        elif isinstance(cell_value, (int, float)):
                            # Números: convertir a string preservando decimales
                            # Para enteros, evitar .0 al final
                            if isinstance(cell_value, float) and cell_value == int(cell_value):
                                cell_value = str(int(cell_value))
                            else:
                                cell_value = str(cell_value)
                        # Strings ya son strings, None se deja como None

                        excel_row_raw[col_name] = cell_value

                case_data = {
                    'caso_prueba': self._get_cell_value(row, headers, ['CASOPRUEBA', 'CASO']),
                    'version': self._get_cell_value(row, headers, ['VERSION']),
                    'encf': self._get_cell_value(row, headers, ['ENCF']),
                    'codigo_seguridad_ecf': self._get_cell_value(row, headers, ['CODIGOSEGURIDADECF', 'CODIGOSEGURIDAD', 'CODSEG']),
                    'sequence': row_idx,
                    'name': f"Caso {row_idx} - Tipo {tipo_ecf}",
                    'tipo_ecf': str(tipo_ecf).strip(),
                    'receptor_rnc': self._get_cell_value(row, headers, ['RNCCOMPRADOR', 'RNCRECEPTOR', 'RNC', 'RECEPTOR RNC', 'RNC RECEPTOR', 'RNC_RECEPTOR']),
                    'receptor_nombre': self._get_cell_value(row, headers, ['RAZONSOCIALCOMPRADOR', 'NOMBRERECEPTOR', 'NOMBRE', 'RECEPTOR', 'NOMBRE RECEPTOR', 'RECEPTOR_NOMBRE']),
                    'fecha_comprobante': self._parse_date(self._get_cell_value(row, headers, ['FECHAEMISION', 'FECHACOMPROBANTE', 'FECHA', 'FECHA COMPROBANTE', 'FECHA_COMPROBANTE'])),
                    'moneda': self._get_cell_value(row, headers, ['MONEDA', 'MONEDAFACTURA', 'CURRENCY', 'TIPOMONEDA']),
                    'tipo_ingreso': self._get_cell_value(row, headers, ['TIPOINGRESOS', 'TIPO INGRESO', 'TIPO_INGRESO', 'INGRESOS']),
                    'tipo_pago': self._get_cell_value(row, headers, ['TIPOPAGO', 'FORMAPAGO[1]', 'FORMAPAGO', 'TIPO PAGO', 'TIPO_PAGO', 'PAGO']),
                    'monto_subtotal': self._parse_float(self._get_cell_value(row, headers, ['MONTOSUBTOTAL', 'SUBTOTAL', 'MONTO SUBTOTAL', 'MONTO_SUBTOTAL'])),
                    'monto_descuento': self._parse_float(self._get_cell_value(row, headers, ['MONTODESCUENTO', 'DESCUENTO'])),
                    'monto_gravado_i1': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADOI1', 'MONTOGRAVADOTASAI1', 'GRAVADO I1', 'MONTO GRAVADO I1', 'GRAVADO_I1'])),
                    'monto_gravado_i2': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADOI2', 'MONTOGRAVADOTASAI2', 'GRAVADO I2', 'MONTO GRAVADO I2', 'GRAVADO_I2'])),
                    'monto_gravado_i3': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADOI3', 'MONTOGRAVADOTASAI3', 'GRAVADO I3', 'MONTO GRAVADO I3', 'GRAVADO_I3'])),
                    'monto_exento': self._parse_float(self._get_cell_value(row, headers, ['MONTOEXENTO', 'EXENTO', 'MONTO EXENTO'])),
                    'total_itbis': self._parse_float(self._get_cell_value(row, headers, ['TOTALITBIS', 'ITBIS', 'TOTAL ITBIS', 'ITBIS TOTAL'])),
                    'total_itbis1': self._parse_float(self._get_cell_value(row, headers, ['ITBIS1', 'ITBIS 1'])),
                    'total_itbis2': self._parse_float(self._get_cell_value(row, headers, ['ITBIS2', 'ITBIS 2'])),
                    'total_itbis3': self._parse_float(self._get_cell_value(row, headers, ['ITBIS3', 'ITBIS 3'])),
                    'monto_total': self._parse_float(self._get_cell_value(row, headers, ['MONTOTOTAL', 'TOTAL', 'MONTO TOTAL', 'MONTO_TOTAL'])),
                    'ncf_modificado': self._get_cell_value(row, headers, ['NCFMODIFICADO', 'ENCFMODIFICADO', 'NCF MODIFICADO', 'NCF_MODIFICADO']),
                    'razon_modificacion': self._map_razon_modificacion(self._get_cell_value(row, headers, ['RAZONMODIFICACION', 'CODIGOMODIFICACION', 'RAZON', 'RAZON MODIFICACION', 'RAZON_MODIFICACION'])),
                    'expected_result': self._get_cell_value(row, headers, ['RESULTADOESPERADO', 'RESULTADO', 'ESPERADO', 'RESULTADO ESPERADO']),
                    'identificador_extranjero': self._get_cell_value(row, headers, ['IDENTIFICADOREXTRANJERO', 'IDEXTRANJERO', 'IDENT_EXTRANJERO']),
                    # Emisor
                    'rnc_emisor': self._get_cell_value(row, headers, ['RNCEMISOR']),
                    'razon_social_emisor': self._get_cell_value(row, headers, ['RAZONSOCIALEMISOR']),
                    'nombre_comercial': self._get_cell_value(row, headers, ['NOMBRECOMERCIAL']),
                    'sucursal': self._get_cell_value(row, headers, ['SUCURSAL']),
                    'direccion_emisor': self._get_cell_value(row, headers, ['DIRECCIONEMISOR']),
                    'municipio_emisor': self._get_cell_value(row, headers, ['MUNICIPIO']),
                    'provincia_emisor': self._get_cell_value(row, headers, ['PROVINCIA']),
                    'correo_emisor': self._get_cell_value(row, headers, ['CORREOEMISOR']),
                    'website_emisor': self._get_cell_value(row, headers, ['WEBSITE']),
                    'actividad_economica': self._get_cell_value(row, headers, ['ACTIVIDADECONOMICA']),
                    'codigo_vendedor': self._get_cell_value(row, headers, ['CODIGOVENDEDOR']),
                    'numero_factura_interna': self._get_cell_value(row, headers, ['NUMEROFACTURAINTERNA']),
                    'numero_pedido_interno': self._get_cell_value(row, headers, ['NUMEROPEDIDOINTERNO']),
                    'zona_venta': self._get_cell_value(row, headers, ['ZONAVENTA']),
                    'ruta_venta': self._get_cell_value(row, headers, ['RUTAVENTA']),
                    'info_adicional_emisor': self._get_cell_value(row, headers, ['INFORMACIONADICIONALEMISOR']),
                    'monto_no_facturable': self._parse_float(self._get_cell_value(row, headers, ['MONTONOFACTURABLE'])),
                }

                # Calcular monto total a pagar
                case_data['monto_total_pagar'] = case_data['monto_total']

                # Calcular precio unitario para la línea
                case_data['cantidad_items'] = 1
                case_data['precio_unitario'] = case_data['monto_subtotal'] or 0
                case_data['descripcion_item'] = f"Prueba e-CF Tipo {tipo_ecf}"
                case_data['pagos'] = self._extract_payments(row, headers)
                case_data['items'] = self._extract_items(row, headers, case_data)

                # ====================================================================
                # CRÍTICO: Guardar fila RAW del Excel para el builder del script
                # El builder necesita los nombres de columna EXACTOS del Excel DGII
                # ====================================================================
                case_data['excel_row_raw'] = excel_row_raw

                cases.append(case_data)

                _logger.debug(f"Caso ECF parseado: {case_data['name']}")

            except Exception as e:
                _logger.warning(f"Error al parsear fila {row_idx} de ECF: {e}")
                continue

        return cases

    def _parse_rfce_sheet(self, sheet):
        """
        Parsea la hoja 'RFCE' del Excel DGII.
        Se soportan dos formatos:
        - Formato antiguo: campos de período/cantidad.
        - Formato observado: columnas detalladas por caso (como las de la solicitud).
        """
        cases = []

        headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                # Normalizar: eliminar espacios, convertir a mayúsculas, eliminar guiones bajos
                normalized_header = str(cell.value).strip().upper().replace(' ', '').replace('_', '')
                headers[normalized_header] = idx

        _logger.info(f"Encabezados encontrados en hoja RFCE: {list(headers.keys())}")

        formato_detallado = 'MONTOTOTAL' in headers and 'FECHAEMISION' in headers

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            try:
                if not formato_detallado:
                    periodo = self._get_cell_value(row, headers, ['MONTOPERIODO', 'PERIODO', 'PERÍODO'])
                    if not periodo:
                        continue

                    case_data = {
                        'caso_prueba': self._get_cell_value(row, headers, ['CASOPRUEBA', 'CASO']),
                        'version': self._get_cell_value(row, headers, ['VERSION']),
                        'encf': self._get_cell_value(row, headers, ['ENCF']),
                        'codigo_seguridad_ecf': self._get_cell_value(row, headers, ['CODIGOSEGURIDADECF', 'CODIGOSEGURIDAD', 'CODSEG']),
                        'sequence': row_idx,
                        'name': f"RFCE {periodo}",
                        'periodo': str(periodo).strip(),
                        'fecha_desde': self._parse_date(self._get_cell_value(row, headers, ['FECHADESDE', 'FECHA DESDE', 'DESDE', 'FECHA_DESDE'])),
                        'fecha_hasta': self._parse_date(self._get_cell_value(row, headers, ['FECHAHASTA', 'FECHA HASTA', 'HASTA', 'FECHA_HASTA'])),
                        'cantidad_comprobantes': int(self._parse_float(self._get_cell_value(row, headers, ['CANTIDADCOMPROBANTES', 'CANTIDAD', 'CANT COMPROBANTES']))),
                        'monto_gravado': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADO', 'GRAVADO', 'MONTO GRAVADO'])),
                        'total_itbis': self._parse_float(self._get_cell_value(row, headers, ['TOTALITBIS', 'ITBIS', 'TOTAL ITBIS'])),
                        'monto_exento': self._parse_float(self._get_cell_value(row, headers, ['MONTOEXENTO', 'EXENTO', 'MONTO EXENTO'])),
                        'monto_total': self._parse_float(self._get_cell_value(row, headers, ['MONTOTOTAL', 'TOTAL', 'MONTO TOTAL'])),
                        'monto_no_facturable': self._parse_float(self._get_cell_value(row, headers, ['MONTONOFACTURABLE'])),
                        'monto_periodo': self._get_cell_value(row, headers, ['MONTOPERIODO', 'PERIODO']),
                        'expected_result': self._get_cell_value(row, headers, ['RESULTADOESPERADO', 'RESULTADO', 'ESPERADO']),
                        'tipo_ingreso': self._get_cell_value(row, headers, ['TIPOINGRESOS', 'TIPO INGRESO', 'TIPO_INGRESO', 'INGRESOS']),
                        'tipo_pago': self._get_cell_value(row, headers, ['TIPOPAGO', 'FORMAPAGO[1]', 'FORMAPAGO', 'TIPO PAGO', 'TIPO_PAGO', 'PAGO']),
                    }
                else:
                    fecha_emision = self._parse_date(self._get_cell_value(row, headers, ['FECHAEMISION']))
                    periodo = fecha_emision.strftime('%Y%m') if fecha_emision else False

                    case_data = {
                        'caso_prueba': self._get_cell_value(row, headers, ['CASOPRUEBA', 'CASO']),
                        'version': self._get_cell_value(row, headers, ['VERSION']),
                        'encf': self._get_cell_value(row, headers, ['ENCF']),
                        'codigo_seguridad_ecf': self._get_cell_value(row, headers, ['CODIGOSEGURIDADECF', 'CODIGOSEGURIDAD', 'CODSEG']),
                        'sequence': row_idx,
                        'name': f"RFCE {periodo or row_idx}",
                        'periodo': periodo,
                        'fecha_desde': fecha_emision,
                        'fecha_hasta': fecha_emision,
                        'cantidad_comprobantes': 1,
                        'monto_gravado': self._parse_float(self._get_cell_value(row, headers, ['MONTOGRAVADOTOTAL', 'MONTOGRAVADO'])),
                        'total_itbis': self._parse_float(self._get_cell_value(row, headers, ['TOTALITBIS'])),
                        'monto_exento': self._parse_float(self._get_cell_value(row, headers, ['MONTOEXENTO'])),
                        'monto_total': self._parse_float(self._get_cell_value(row, headers, ['MONTOTOTAL'])),
                        'monto_no_facturable': self._parse_float(self._get_cell_value(row, headers, ['MONTONOFACTURABLE'])),
                        'monto_periodo': self._get_cell_value(row, headers, ['MONTOPERIODO']),
                        'expected_result': None,
                        'tipo_ingreso': self._get_cell_value(row, headers, ['TIPOINGRESOS', 'TIPO INGRESO', 'TIPO_INGRESO', 'INGRESOS']),
                        'tipo_pago': self._get_cell_value(row, headers, ['TIPOPAGO', 'FORMAPAGO[1]', 'FORMAPAGO', 'TIPO PAGO', 'TIPO_PAGO', 'PAGO']),
                    }

                case_data['pagos'] = self._extract_payments(row, headers)
                case_data['impuestos_adicionales'] = self._extract_impuestos_adicionales(row, headers)

                cases.append(case_data)
                _logger.debug(f"Caso RFCE parseado: {case_data['name']}")

            except Exception as e:
                _logger.warning(f"Error al parsear fila {row_idx} de RFCE: {e}")
                continue

        return cases

    def _get_cell_value(self, row, headers, possible_names):
        """Obtiene el valor de una celda buscando en múltiples nombres posibles"""
        for name in possible_names:
            # Normalizar el nombre de búsqueda de la misma forma que los headers
            normalized_name = name.upper().replace(' ', '').replace('_', '')
            if normalized_name in headers:
                col_idx = headers[normalized_name] - 1
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        # Filtrar errores de Excel como #e, #N/A, #VALUE!, etc.
                        if isinstance(value, str) and value.strip().startswith('#'):
                            return None
                        return value
        return None

    def _parse_float(self, value):
        """Convierte un valor a float de manera segura"""
        if value is None or value == '':
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _parse_date(self, value):
        """Convierte un valor a fecha de manera segura"""
        return self._normalize_date(value)

    def _map_razon_modificacion(self, value):
        """
        Mapea descripciones textuales a códigos de razón de modificación.
        Según normativa DGII para e-CF.
        """
        if not value or (isinstance(value, str) and value.strip().startswith('#')):
            return None

        value_str = str(value).strip().lower()

        # Mapeo de descripciones textuales a códigos
        mapping = {
            # Códigos directos (si ya viene el código)
            '01': '01',
            '02': '02',
            '03': '03',
            '04': '04',
            '05': '05',

            # Descripciones textuales comunes
            'anulacion': '01',
            'anulación': '01',
            'cancelacion': '01',
            'cancelación': '01',

            'correccion': '02',
            'corrección': '02',
            'error': '02',
            'error en monto': '02',
            'error de monto': '02',
            'error en datos': '02',
            'error en precio': '02',

            'devolucion': '03',
            'devolución': '03',
            'retorno': '03',

            'descuento': '04',
            'rebaja': '04',

            'bonificacion': '05',
            'bonificación': '05',
        }

        return mapping.get(value_str, None)

    def run_tests(self):
        """
        Lee el archivo Excel o CSV DGII y crea el set de pruebas con todos los casos
        """
        self.ensure_one()

        # Detectar si es CSV o Excel
        is_csv = self._is_csv_file()

        ecf_cases_data = []
        rfce_cases_data = []

        if is_csv:
            _logger.info("Procesando archivo CSV DGII")
            csv_rows = self._parse_csv_to_sheet()

            # Para CSV, asumimos que todos los datos están en un solo archivo
            # El CSV de DGII típicamente contiene solo casos e-CF
            ecf_cases_data = self._parse_ecf_csv(csv_rows)

        else:
            _logger.info("Procesando archivo Excel DGII")
            workbook = self._get_workbook()

            # Buscar hojas ECF y RFCE
            ecf_sheet = None
            rfce_sheet = None

            for sheet_name in workbook.sheetnames:
                sheet_name_upper = sheet_name.upper()
                # Importante: RFCE contiene "ECF" como substring, por eso se evalúa primero RFCE.
                if 'RFCE' in sheet_name_upper or 'RESUMEN' in sheet_name_upper:
                    rfce_sheet = workbook[sheet_name]
                    _logger.info(f"Hoja RFCE encontrada: {sheet_name}")
                elif 'ECF' in sheet_name_upper:
                    ecf_sheet = workbook[sheet_name]
                    _logger.info(f"Hoja ECF encontrada: {sheet_name}")

            if not ecf_sheet and not rfce_sheet:
                raise UserError(_("No se encontraron hojas ECF o RFCE en el archivo Excel."))

            # Parsear hojas de Excel
            if ecf_sheet:
                ecf_cases_data = self._parse_ecf_sheet(ecf_sheet)
            if rfce_sheet:
                rfce_cases_data = self._parse_rfce_sheet(rfce_sheet)

        if not ecf_cases_data and not rfce_cases_data:
            raise UserError(_("No se encontraron casos en el archivo."))

        # Crear el set de pruebas
        file_type = "CSV" if is_csv else "Excel"
        test_set = self.env['ecf.test.set'].create({
            'name': self.test_set_name,
            'description': f'Importado desde {self.filename or f"archivo {file_type}"} el {fields.Datetime.now()}'
        })

        id_lote = str(uuid.uuid4())
        ecf_cases_created = 0
        rfce_cases_created = 0

        for case_data in ecf_cases_data:
            _logger.info(f"[IMPORT] Procesando caso fila {case_data.get('sequence')} - tipo {case_data.get('tipo_ecf')}")

            # Verificar si tenemos excel_row_raw
            has_raw = bool(case_data.get('excel_row_raw'))
            if has_raw:
                _logger.info(f"[IMPORT] Fila RAW disponible con {len(case_data.get('excel_row_raw', {}))} columnas")
            else:
                _logger.warning(f"[IMPORT] NO hay fila RAW para caso {case_data.get('sequence')}")

            case_vals = {
                'test_set_id': test_set.id,
                'sequence': case_data.get('sequence'),
                'name': case_data.get('name'),
                'tipo_ecf': case_data.get('tipo_ecf'),
                'receptor_rnc': case_data.get('receptor_rnc'),
                'receptor_nombre': case_data.get('receptor_nombre'),
                'receptor_tipo_identificacion': case_data.get('receptor_tipo_identificacion') or '1',
                'identificador_extranjero': case_data.get('identificador_extranjero'),
                'fecha_comprobante': case_data.get('fecha_comprobante'),
                'moneda': case_data.get('moneda') or "DOP",
                'tipo_ingreso': case_data.get('tipo_ingreso'),
                'tipo_pago': case_data.get('tipo_pago'),
                'monto_subtotal': case_data.get('monto_subtotal'),
                'monto_descuento': case_data.get('monto_descuento'),
                'monto_gravado_i1': case_data.get('monto_gravado_i1'),
                'monto_gravado_i2': case_data.get('monto_gravado_i2'),
                'monto_gravado_i3': case_data.get('monto_gravado_i3'),
                'monto_exento': case_data.get('monto_exento'),
                'total_itbis': case_data.get('total_itbis'),
                'total_itbis1': case_data.get('total_itbis1'),
                'total_itbis2': case_data.get('total_itbis2'),
                'total_itbis3': case_data.get('total_itbis3'),
                'monto_total': case_data.get('monto_total'),
                'monto_total_pagar': case_data.get('monto_total'),
                'ncf_modificado': case_data.get('ncf_modificado'),
                'razon_modificacion': case_data.get('razon_modificacion'),
                'cantidad_items': case_data.get('cantidad_items') or 1,
                'descripcion_item': case_data.get('descripcion_item'),
                'precio_unitario': case_data.get('precio_unitario'),
                'expected_result': case_data.get('expected_result'),
            }

            case = self.env['ecf.test.case'].create(case_vals)
            _logger.info(f"[IMPORT] Caso creado con ID {case.id}")

            try:
                # Validar datos (validación permisiva - solo tipo_ecf requerido)
                es_valido, errores_validacion = self._validate_case_data(case_data)

                if not es_valido:
                    # Marcar caso como error de validación
                    error_msg = "Errores de validación: " + "; ".join(errores_validacion)
                    _logger.warning(f"[IMPORT] Fila {case_data.get('sequence')}: {error_msg}")
                    case.mark_error(error_msg)
                    continue

                # Construir payload usando el builder del script probado
                _logger.info(f"[IMPORT] Construyendo JSON para caso {case.id}...")
                payload, hash_input = self._build_canonical_payload(case_data, id_lote)

                _logger.info(f"[IMPORT] JSON construido, guardando en caso {case.id}...")
                case.set_payload(payload, hash_input, id_lote, case_data.get('sequence'))

                _logger.info(f"[IMPORT] Caso {case.id} procesado exitosamente - estado: {case.state}")
                ecf_cases_created += 1

            except Exception as e:
                _logger.error(f"[IMPORT] Error al procesar fila {case_data.get('sequence')}: {str(e)}", exc_info=True)
                case.mark_error(str(e))
                continue

        _logger.info(f"Set de pruebas creado: {ecf_cases_created} casos ECF (id_lote {id_lote})")

        # Crear casos RFCE (solo registro, sin envío a API)
        for case_data in rfce_cases_data:
            # Garantizar campos requeridos para el modelo
            if not case_data.get('fecha_desde'):
                case_data['fecha_desde'] = case_data.get('fecha_hasta') or fields.Date.today()
            if not case_data.get('fecha_hasta'):
                case_data['fecha_hasta'] = case_data.get('fecha_desde') or fields.Date.today()

            # Filtrar solo los campos que existen en el modelo ecf.test.rfce.case
            rfce_vals = {
                'test_set_id': test_set.id,
                'sequence': case_data.get('sequence'),
                'name': case_data.get('name'),
                'caso_prueba': case_data.get('caso_prueba'),
                'version': case_data.get('version'),
                'encf': case_data.get('encf'),
                'codigo_seguridad_ecf': case_data.get('codigo_seguridad_ecf'),
                'fecha_desde': case_data.get('fecha_desde'),
                'fecha_hasta': case_data.get('fecha_hasta'),
                'periodo': case_data.get('periodo'),
                'cantidad_comprobantes': case_data.get('cantidad_comprobantes') or 0,
                'monto_total': case_data.get('monto_total') or 0.0,
                'monto_gravado': case_data.get('monto_gravado') or 0.0,
                'total_itbis': case_data.get('total_itbis') or 0.0,
                'monto_exento': case_data.get('monto_exento') or 0.0,
                'monto_no_facturable': case_data.get('monto_no_facturable') or 0.0,
                'monto_periodo': case_data.get('monto_periodo'),
                'tipo_ingreso': case_data.get('tipo_ingreso'),
                'tipo_pago': case_data.get('tipo_pago'),
                'expected_result': case_data.get('expected_result'),
            }

            rfce_case = self.env['ecf.test.rfce.case'].create(rfce_vals)
            try:
                payload_rfce, hash_rfce = self._build_rfce_payload(case_data, id_lote)
                rfce_case.set_payload(payload_rfce, hash_rfce, id_lote, case_data.get('sequence'))
            except Exception as e:
                _logger.warning("No se pudo construir payload RFCE para fila %s: %s", case_data.get('sequence'), e)
            rfce_cases_created += 1

        if rfce_cases_created:
            _logger.info(f"Casos RFCE registrados: {rfce_cases_created}")

        # ========================================================================
        # ENVÍO A API usando sistema de proveedores
        # ========================================================================
        if self.send_to_api:
            # Obtener el proveedor de API por defecto
            provider = self.env['ecf.api.provider'].get_default_provider()

            if not provider:
                raise UserError(_(
                    "No hay proveedor de API configurado.\n"
                    "Configure uno en: e-CF Tests > Proveedores de API"
                ))

            _logger.info(f"Usando proveedor de API: {provider.name} ({provider.provider_type})")

            # Obtener casos del set que tienen payload listo
            cases_to_send = test_set.ecf_case_ids.filtered(lambda c: c.state == 'payload_ready')

            # Aplicar filtro por tipo si está especificado
            if self.filter_tipo_ecf and self.filter_tipo_ecf != 'all':
                cases_to_send = cases_to_send.filtered(lambda c: c.tipo_ecf == self.filter_tipo_ecf)
                _logger.info(f"Filtrado por tipo {self.filter_tipo_ecf}: {len(cases_to_send)} casos")

            if not cases_to_send:
                _logger.warning("No hay casos con payload listo para enviar (después de filtros)")
            else:
                # Ordenar casos según prioridad DGII
                cases_with_priority = []
                for case in cases_to_send:
                    try:
                        doc = json.loads(case.payload_json)
                        priority = self._get_tipo_ecf_priority(doc)
                        cases_with_priority.append((priority, case, doc))
                    except Exception as e:
                        _logger.error(f"Error al parsear payload del caso {case.id}: {e}")
                        continue

                # Ordenar por prioridad
                cases_with_priority.sort(key=lambda x: x[0])

                _logger.info(f"Enviando {len(cases_with_priority)} casos via {provider.name} en orden DGII")

                # Enviar documentos ordenados
                ok = 0
                for priority, case, doc in cases_with_priority:
                    tipo_ecf = doc["ECF"]["Encabezado"]["IdDoc"].get("TipoeCF", "??")
                    encf = doc["ECF"]["Encabezado"]["IdDoc"].get("eNCF", "N/A")
                    rnc = doc["ECF"]["Encabezado"]["Emisor"].get("RNCEmisor")

                    try:
                        # Enviar usando el proveedor (registra en log automáticamente)
                        success, resp_data, track_id, error_msg, raw_response, signed_xml = provider.send_ecf(
                            doc, rnc=rnc, encf=encf,
                            origin='test_set',
                            test_case_id=case.id
                        )

                        resp_text = json.dumps(resp_data, ensure_ascii=False) if isinstance(resp_data, dict) else str(resp_data or error_msg)

                        if success:
                            ok += 1
                            case.mark_sent(resp_text, track_id=track_id, accepted=True, rejected=False,
                                           raw_response=raw_response, signed_xml=signed_xml)
                            _logger.info(f"Enviado caso {case.id} - Tipo {tipo_ecf} - eNCF {encf} - ACEPTADO")
                        else:
                            case.mark_sent(resp_text, track_id=track_id, accepted=False, rejected=True,
                                           raw_response=raw_response, signed_xml=signed_xml)
                            _logger.info(f"Enviado caso {case.id} - Tipo {tipo_ecf} - eNCF {encf} - RECHAZADO: {error_msg}")

                    except Exception as e:
                        error_msg = f"Error al enviar via {provider.name}: {str(e)}"
                        _logger.error(f"Caso {case.id}: {error_msg}", exc_info=True)
                        case.mark_error(error_msg)

                _logger.info(f"Resultado final: {ok}/{len(cases_with_priority)} casos aceptados via {provider.name}")

        test_set.write({'state': 'completed'})

        # Abrir el set de pruebas creado
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ecf.test.set',
            'res_id': test_set.id,
            'view_mode': 'form',
            'target': 'current',
            'context': {
                'form_view_initial_mode': 'edit',
            }
        }
